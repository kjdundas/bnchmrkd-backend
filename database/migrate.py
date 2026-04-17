"""
bnchmrkd. Data Migration Script
================================
Loads all data from Olympic_Sprints_Master_Database_v5.xlsx into Supabase PostgreSQL.

Usage:
    pip install openpyxl psycopg2-binary
    python migrate.py --db-url "postgresql://postgres:PASSWORD@db.XXXXX.supabase.co:5432/postgres"

Or set the DATABASE_URL environment variable.
"""

import argparse
import os
import sys
import time
from datetime import datetime, date

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ============================================================
# CONFIG
# ============================================================

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Olympic_Sprints_Master_Database_v5.xlsx")

# Map Excel discipline names + gender to our discipline codes
DISCIPLINE_MAP = {
    ("100 Metres", "Male"): "M100",
    ("100 Metres", "Female"): "F100",
    ("200 Metres", "Male"): "M200",
    ("200 Metres", "Female"): "F200",
    ("400 Metres", "Male"): "M400",
    ("400 Metres", "Female"): "F400",
    ("110 Metres Hurdles", "Male"): "M110H",
    ("100 Metres Hurdles", "Female"): "F100H",
    ("400 Metres Hurdles", "Male"): "M400H",
    ("400 Metres Hurdles", "Female"): "F400H",
}

# Fallback mappings for variations in the data
DISCIPLINE_NAME_NORMALIZE = {
    "100m": "100 Metres",
    "200m": "200 Metres",
    "400m": "400 Metres",
    "110mh": "110 Metres Hurdles",
    "100mh": "100 Metres Hurdles",
    "400mh": "400 Metres Hurdles",
    "100 metres": "100 Metres",
    "200 metres": "200 Metres",
    "400 metres": "400 Metres",
    "110 metres hurdles": "110 Metres Hurdles",
    "100 metres hurdles": "100 Metres Hurdles",
    "400 metres hurdles": "400 Metres Hurdles",
}

GENDER_MAP = {
    "Male": "M",
    "Female": "F",
    "M": "M",
    "F": "F",
}


def normalize_discipline(name):
    if not name:
        return name
    return DISCIPLINE_NAME_NORMALIZE.get(name.lower().strip(), name.strip())


def get_discipline_code(discipline_name, gender):
    disc = normalize_discipline(discipline_name)
    g = gender.strip() if gender else ""
    # Normalize gender
    if g in ("M", "Male"):
        g = "Male"
    elif g in ("F", "Female"):
        g = "Female"
    return DISCIPLINE_MAP.get((disc, g))


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def to_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def to_bool(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.lower() in ("true", "1", "yes")
    return None


# ============================================================
# MIGRATION
# ============================================================

def run_migration(db_url):
    print(f"Connecting to database...")
    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    cur = conn.cursor()

    print(f"Loading Excel workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    try:
        # ----------------------------------------------------------
        # Step 1: Load discipline ID lookup
        # ----------------------------------------------------------
        cur.execute("SELECT id, code FROM disciplines")
        disc_lookup = {row[1]: row[0] for row in cur.fetchall()}
        print(f"  Disciplines in DB: {list(disc_lookup.keys())}")

        # ----------------------------------------------------------
        # Step 2: Load Athletes from "Athlete Summary" sheet
        # ----------------------------------------------------------
        print("\n--- Loading Athletes ---")
        ws = wb["Athlete Summary"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"  {len(rows)} rows in Athlete Summary")

        athlete_lookup = {}  # name -> athlete_id

        athlete_data = []
        for row in rows:
            name = row[0]
            if not name:
                continue
            gender_raw = row[1]
            gender = GENDER_MAP.get(gender_raw.strip(), "M") if gender_raw else "M"
            dob = to_date(row[2])
            country = row[3]
            doping = row[5] if len(row) > 5 else None

            athlete_data.append((name.strip(), gender, country, dob, True, doping))

        # Bulk insert athletes
        execute_values(
            cur,
            """INSERT INTO athletes (name, gender, country, date_of_birth, is_olympic, doping_flag)
               VALUES %s
               ON CONFLICT DO NOTHING
               RETURNING id, name""",
            athlete_data,
            template="(%s, %s, %s, %s, %s, %s)"
        )

        conn.commit()

        # Build lookup
        cur.execute("SELECT id, name FROM athletes")
        for row in cur.fetchall():
            athlete_lookup[row[1]] = row[0]

        print(f"  Loaded {len(athlete_lookup)} athletes")

        # ----------------------------------------------------------
        # Step 3: Load Personal Bests
        # ----------------------------------------------------------
        print("\n--- Loading Personal Bests ---")
        ws = wb["Personal Bests"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"  {len(rows)} rows in Personal Bests")

        pb_data = []
        skipped = 0
        for row in rows:
            name = row[0]
            if not name:
                continue
            name = name.strip()
            athlete_id = athlete_lookup.get(name)
            if not athlete_id:
                skipped += 1
                continue

            disc_name = row[1]
            gender_raw = row[2]
            disc_code = get_discipline_code(disc_name, gender_raw)
            if not disc_code or disc_code not in disc_lookup:
                skipped += 1
                continue

            disc_id = disc_lookup[disc_code]
            pb_time = to_float(row[3])
            if not pb_time:
                skipped += 1
                continue

            pb_date = to_date(row[4])
            total_races = to_int(row[5])
            avg_time = to_float(row[6])
            median_time = to_float(row[7])
            std_dev = to_float(row[8])

            pb_data.append((athlete_id, disc_id, pb_time, pb_date, total_races, avg_time, median_time, std_dev))

        execute_values(
            cur,
            """INSERT INTO personal_bests (athlete_id, discipline_id, pb_time, pb_date, total_races, avg_time, median_time, std_dev)
               VALUES %s
               ON CONFLICT (athlete_id, discipline_id) DO UPDATE SET
                   pb_time = EXCLUDED.pb_time,
                   pb_date = EXCLUDED.pb_date,
                   total_races = EXCLUDED.total_races""",
            pb_data,
            template="(%s, %s, %s, %s, %s, %s, %s, %s)"
        )
        conn.commit()
        print(f"  Loaded {len(pb_data)} personal bests (skipped {skipped})")

        # ----------------------------------------------------------
        # Step 4: Load Career Sprints (311K rows — batch insert)
        # ----------------------------------------------------------
        print("\n--- Loading Career Sprints ---")
        ws = wb["Career Sprints"]

        BATCH_SIZE = 5000
        total_loaded = 0
        total_skipped = 0
        batch = []

        row_iter = ws.iter_rows(min_row=2, values_only=True)
        start_time = time.time()

        for row in row_iter:
            athlete_name = row[12]
            if not athlete_name:
                total_skipped += 1
                continue
            athlete_name = athlete_name.strip()

            athlete_id = athlete_lookup.get(athlete_name)
            if not athlete_id:
                total_skipped += 1
                continue

            disc_name = row[1]
            gender_raw = row[17]
            disc_code = get_discipline_code(disc_name, gender_raw)
            if not disc_code or disc_code not in disc_lookup:
                total_skipped += 1
                continue

            disc_id = disc_lookup[disc_code]
            time_seconds = to_float(row[2])
            if not time_seconds or time_seconds <= 0:
                total_skipped += 1
                continue

            race_date = to_date(row[0])
            wind_mps = to_float(row[4])
            wind_legal = to_bool(row[5])
            competition = str(row[6])[:300] if row[6] else None
            country = str(row[7])[:150] if row[7] else None
            category = str(row[8])[:50] if row[8] else None
            race_type = str(row[9])[:50] if row[9] else None
            place = str(row[10])[:10] if row[10] else None
            score = to_int(row[11])
            age_decimal = to_float(row[14])
            age_years = to_int(row[15])
            is_olympic = to_bool(row[16])
            season_year = to_int(row[13])

            batch.append((
                athlete_id, disc_id, race_date, time_seconds,
                wind_mps, wind_legal, competition, country,
                category, race_type, place, score,
                age_decimal, age_years, is_olympic, season_year
            ))

            if len(batch) >= BATCH_SIZE:
                execute_values(
                    cur,
                    """INSERT INTO race_results
                       (athlete_id, discipline_id, race_date, time_seconds,
                        wind_mps, wind_legal, competition, country,
                        category, race_type, place, score,
                        age_decimal, age_years, is_olympic_race, season_year)
                       VALUES %s""",
                    batch,
                    template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                )
                conn.commit()
                total_loaded += len(batch)
                elapsed = time.time() - start_time
                rate = total_loaded / elapsed if elapsed > 0 else 0
                print(f"  {total_loaded:,} rows loaded ({rate:.0f} rows/sec)...", end="\r")
                batch = []

        # Final batch
        if batch:
            execute_values(
                cur,
                """INSERT INTO race_results
                   (athlete_id, discipline_id, race_date, time_seconds,
                    wind_mps, wind_legal, competition, country,
                    category, race_type, place, score,
                    age_decimal, age_years, is_olympic_race, season_year)
                   VALUES %s""",
                batch,
                template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            )
            conn.commit()
            total_loaded += len(batch)

        elapsed = time.time() - start_time
        print(f"\n  Loaded {total_loaded:,} race results in {elapsed:.1f}s (skipped {total_skipped:,})")

        # ----------------------------------------------------------
        # Step 5: Load Olympic Results
        # ----------------------------------------------------------
        print("\n--- Loading Olympic Results ---")
        ws = wb["Olympic Results"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"  {len(rows)} rows in Olympic Results")

        oly_data = []
        skipped = 0
        for row in rows:
            disc_name = row[0]
            gender_raw = row[14]
            disc_code = get_discipline_code(disc_name, gender_raw)
            if not disc_code or disc_code not in disc_lookup:
                skipped += 1
                continue

            athlete_name = row[7]
            if not athlete_name:
                skipped += 1
                continue
            athlete_name = athlete_name.strip()
            athlete_id = athlete_lookup.get(athlete_name)
            if not athlete_id:
                skipped += 1
                continue

            disc_id = disc_lookup[disc_code]
            games = row[1]
            year = to_int(row[2])
            round_name = row[3]
            heat = row[4]
            rank = to_int(row[5])
            lane = to_int(row[6])
            time_seconds = to_float(row[9])
            reaction_time = to_float(row[10])
            status = row[11]
            qualified = row[12]
            notes = row[13]

            oly_data.append((
                athlete_id, disc_id, games, year, round_name, heat,
                rank, lane, time_seconds, reaction_time, status, qualified, notes
            ))

        execute_values(
            cur,
            """INSERT INTO olympic_results
               (athlete_id, discipline_id, games, year, round, heat,
                rank, lane, time_seconds, reaction_time, status, qualified, notes)
               VALUES %s""",
            oly_data,
            template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        )
        conn.commit()
        print(f"  Loaded {len(oly_data)} Olympic results (skipped {skipped})")

        # ----------------------------------------------------------
        # Step 6: Populate season_bests
        # ----------------------------------------------------------
        print("\n--- Populating Season Bests ---")
        cur.execute("SELECT populate_season_bests()")
        conn.commit()
        cur.execute("SELECT COUNT(*) FROM season_bests")
        count = cur.fetchone()[0]
        print(f"  Generated {count:,} season bests")

        # ----------------------------------------------------------
        # Step 7: Final stats
        # ----------------------------------------------------------
        print("\n--- Migration Complete ---")
        for table in ["athletes", "race_results", "personal_bests", "olympic_results", "season_bests"]:
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            count = cur.fetchone()[0]
            print(f"  {table}: {count:,} rows")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        wb.close()
        cur.close()
        conn.close()
        print("\nDone.")


# ============================================================
# BENCHMARK DATA MIGRATION
# ============================================================

def migrate_benchmarks(db_url):
    """Load benchmark data from the Python constants into the database."""
    print("\n--- Loading Benchmark Data ---")

    # Import benchmark data from the backend
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "bnchmarkd-app", "backend"))
    from app.core.benchmarks import (
        AGE_PERFORMANCE_PERCENTILES,
        ROC_THRESHOLDS,
        TRAJECTORY_CENTROIDS,
        MODEL_COEFFICIENTS,
        IMPROVEMENT_NORMS,
        MODEL_CALIBRATION,
    )

    conn = psycopg2.connect(db_url)
    cur = conn.cursor()

    try:
        # Discipline lookup
        cur.execute("SELECT id, code FROM disciplines")
        disc_lookup = {row[1]: row[0] for row in cur.fetchall()}

        # Age percentiles
        percentile_rows = []
        for disc_code, ages in AGE_PERFORMANCE_PERCENTILES.items():
            disc_id = disc_lookup.get(disc_code)
            if not disc_id:
                continue
            for age, perc in ages.items():
                percentile_rows.append((disc_id, age, perc.p10, perc.p25, perc.p50, perc.p75, perc.p90))

        execute_values(
            cur,
            """INSERT INTO age_percentile_benchmarks (discipline_id, age, p10, p25, p50, p75, p90)
               VALUES %s ON CONFLICT (discipline_id, age) DO UPDATE SET
               p10=EXCLUDED.p10, p25=EXCLUDED.p25, p50=EXCLUDED.p50, p75=EXCLUDED.p75, p90=EXCLUDED.p90""",
            percentile_rows,
            template="(%s, %s, %s, %s, %s, %s, %s)"
        )
        print(f"  age_percentile_benchmarks: {len(percentile_rows)} rows")

        # ROC thresholds
        roc_rows = []
        for disc_code, roc in ROC_THRESHOLDS.items():
            disc_id = disc_lookup.get(disc_code)
            if not disc_id:
                continue
            roc_rows.append((disc_id, roc.optimal_threshold, roc.threshold_90_sensitivity,
                           roc.threshold_80_sensitivity, roc.threshold_70_sensitivity))

        execute_values(
            cur,
            """INSERT INTO roc_thresholds (discipline_id, optimal_threshold, threshold_90_sensitivity,
               threshold_80_sensitivity, threshold_70_sensitivity)
               VALUES %s ON CONFLICT (discipline_id) DO UPDATE SET
               optimal_threshold=EXCLUDED.optimal_threshold""",
            roc_rows,
            template="(%s, %s, %s, %s, %s)"
        )
        print(f"  roc_thresholds: {len(roc_rows)} rows")

        # Trajectory clusters
        cluster_rows = []
        for disc_code, clusters in TRAJECTORY_CENTROIDS.items():
            disc_id = disc_lookup.get(disc_code)
            if not disc_id:
                continue
            for idx, cluster in enumerate(clusters):
                cluster_rows.append((disc_id, idx, cluster.name, cluster.description, cluster.pct_off_pb))

        execute_values(
            cur,
            """INSERT INTO trajectory_clusters (discipline_id, cluster_index, cluster_name, description, centroid_values)
               VALUES %s ON CONFLICT (discipline_id, cluster_index) DO UPDATE SET
               cluster_name=EXCLUDED.cluster_name, description=EXCLUDED.description,
               centroid_values=EXCLUDED.centroid_values""",
            cluster_rows,
            template="(%s, %s, %s, %s, %s)"
        )
        print(f"  trajectory_clusters: {len(cluster_rows)} rows")

        # Model coefficients
        coeff = MODEL_COEFFICIENTS
        coeff_rows = [
            ("best_18_20_z", coeff.best_18_20_z),
            ("pct_rank_at_20", coeff.pct_rank_at_20),
            ("improvement_y0_y2_z", coeff.improvement_y0_y2_z),
            ("consistency_std_z", coeff.consistency_std_z),
            ("races_18_20", coeff.races_18_20),
            ("intercept", coeff.intercept),
        ]
        execute_values(
            cur,
            """INSERT INTO model_coefficients (coefficient_name, coefficient_value)
               VALUES %s ON CONFLICT (coefficient_name) DO UPDATE SET
               coefficient_value=EXCLUDED.coefficient_value""",
            coeff_rows,
            template="(%s, %s)"
        )
        print(f"  model_coefficients: {len(coeff_rows)} rows")

        # Improvement norms
        norm_rows = []
        for disc_code, norms in IMPROVEMENT_NORMS.items():
            disc_id = disc_lookup.get(disc_code)
            if not disc_id:
                continue
            norm_rows.append((disc_id, norms.finalist_median_pct, norms.finalist_std_pct,
                            norms.non_finalist_median_pct, norms.non_finalist_std_pct))

        execute_values(
            cur,
            """INSERT INTO improvement_norms (discipline_id, finalist_median_pct, finalist_std_pct,
               non_finalist_median_pct, non_finalist_std_pct)
               VALUES %s ON CONFLICT (discipline_id) DO UPDATE SET
               finalist_median_pct=EXCLUDED.finalist_median_pct""",
            norm_rows,
            template="(%s, %s, %s, %s, %s)"
        )
        print(f"  improvement_norms: {len(norm_rows)} rows")

        # Model calibration
        cal_rows = []
        for disc_code, cal in MODEL_CALIBRATION.items():
            disc_id = disc_lookup.get(disc_code)
            if not disc_id:
                continue
            cal_rows.append((disc_id, cal.mean_time, cal.std_time))

        execute_values(
            cur,
            """INSERT INTO model_calibration (discipline_id, mean_time, std_time)
               VALUES %s ON CONFLICT (discipline_id) DO UPDATE SET
               mean_time=EXCLUDED.mean_time, std_time=EXCLUDED.std_time""",
            cal_rows,
            template="(%s, %s, %s)"
        )
        print(f"  model_calibration: {len(cal_rows)} rows")

        conn.commit()
        print("  Benchmark data loaded successfully!")

    except Exception as e:
        conn.rollback()
        print(f"ERROR loading benchmarks: {e}")
        raise
    finally:
        cur.close()
        conn.close()


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate bnchmrkd data to Supabase PostgreSQL")
    parser.add_argument("--db-url", type=str, help="PostgreSQL connection URL")
    parser.add_argument("--benchmarks-only", action="store_true", help="Only load benchmark data")
    args = parser.parse_args()

    db_url = args.db_url or os.environ.get("DATABASE_URL")
    if not db_url:
        print("ERROR: Provide --db-url or set DATABASE_URL environment variable")
        print('  Example: python migrate.py --db-url "postgresql://postgres:PASSWORD@db.XXXXX.supabase.co:5432/postgres"')
        sys.exit(1)

    if args.benchmarks_only:
        migrate_benchmarks(db_url)
    else:
        run_migration(db_url)
        migrate_benchmarks(db_url)
