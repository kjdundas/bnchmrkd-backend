"""
bnchmrkd. Throws & Jumps Migration Script (REST API version)
=============================================================
Uses Supabase REST API (PostgREST) instead of direct PostgreSQL connection.
Works around IPv4/pooler connection issues on free tier.

Usage:
    python migrate_throws_jumps_rest.py --supabase-url "https://xxx.supabase.co" --service-key "sb_secret_xxx"

    # Throws only or jumps only:
    python migrate_throws_jumps_rest.py ... --throws-only
    python migrate_throws_jumps_rest.py ... --jumps-only
"""

import argparse
import csv
import json
import math
import os
import re
import sys
import time
import urllib.request
import urllib.error
import urllib.parse

import openpyxl


# ============================================================
# CONFIG
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
THROWS_EXCEL = os.path.join(BASE_DIR, "..", "Olympic_Throws_Master_Database.xlsx")
JUMPS_DIR = os.path.join(BASE_DIR, "..", "jumps_data")

BATCH_SIZE = 500  # REST API batch size (smaller than SQL batches)

# ── Discipline mappings ──

THROWS_DISCIPLINE_MAP = {
    ("Discus Throw", "Male"): "MDT", ("Discus Throw", "Female"): "FDT",
    ("Javelin Throw", "Male"): "MJT", ("Javelin Throw", "Female"): "FJT",
    ("Hammer Throw", "Male"): "MHT", ("Hammer Throw", "Female"): "FHT",
    ("Shot Put", "Male"): "MSP", ("Shot Put", "Female"): "FSP",
}

JUMPS_DISCIPLINE_MAP = {
    ("High Jump", "Male"): "MHJ", ("High Jump", "Female"): "FHJ",
    ("Long Jump", "Male"): "MLJ", ("Long Jump", "Female"): "FLJ",
    ("Triple Jump", "Male"): "MTJ", ("Triple Jump", "Female"): "FTJ",
    ("Pole Vault", "Male"): "MPV", ("Pole Vault", "Female"): "FPV",
}

DISCIPLINE_NORMALIZE = {
    "discus throw": "Discus Throw", "javelin throw": "Javelin Throw",
    "hammer throw": "Hammer Throw", "shot put": "Shot Put",
    "high jump": "High Jump", "long jump": "Long Jump",
    "triple jump": "Triple Jump", "pole vault": "Pole Vault",
}

GENDER_MAP = {"Male": "Male", "Female": "Female", "M": "Male", "F": "Female",
              "male": "Male", "female": "Female"}

JUMPS_DISCIPLINES_INSERT = [
    {"code": "MHJ", "name": "High Jump",   "gender": "M", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": False},
    {"code": "FHJ", "name": "High Jump",   "gender": "F", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": False},
    {"code": "MLJ", "name": "Long Jump",   "gender": "M", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": True},
    {"code": "FLJ", "name": "Long Jump",   "gender": "F", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": True},
    {"code": "MTJ", "name": "Triple Jump",  "gender": "M", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": True},
    {"code": "FTJ", "name": "Triple Jump",  "gender": "F", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": True},
    {"code": "MPV", "name": "Pole Vault",   "gender": "M", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": False},
    {"code": "FPV", "name": "Pole Vault",   "gender": "F", "distance_m": 0, "is_hurdles": False, "is_throws": False, "direction": "desc", "wind_applicable": False},
]

JUMPS_CSV_MAP = {
    "High_Jump_Men_WA_Results.csv": ("High Jump", "Male"),
    "High_Jump_Women_WA_Results.csv": ("High Jump", "Female"),
    "Long_Jump_Men_WA_Results.csv": ("Long Jump", "Male"),
    "Long_Jump_Women_WA_Results.csv": ("Long Jump", "Female"),
    "Pole_Vault_Men_WA_Results.csv": ("Pole Vault", "Male"),
    "Pole_Vault_Women_WA_Results.csv": ("Pole Vault", "Female"),
    "Triple_Jump_Men_WA_Results.csv": ("Triple Jump", "Male"),
    "Triple_Jump_Women_WA_Results.csv": ("Triple Jump", "Female"),
}


# ============================================================
# REST API HELPER
# ============================================================

class SupabaseREST:
    def __init__(self, url, service_key):
        self.base_url = url.rstrip("/")
        self.service_key = service_key
        self.headers = {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }

    def get(self, table, params=""):
        """GET from a table with optional query params. Auto-paginates for large tables."""
        all_rows = []
        page_size = 1000
        offset = 0
        while True:
            sep = "&" if params else ""
            url = f"{self.base_url}/rest/v1/{table}?{params}{sep}limit={page_size}&offset={offset}"
            req = urllib.request.Request(url, headers={
                **self.headers,
                "Prefer": "return=representation",
            })
            resp = urllib.request.urlopen(req, timeout=30)
            rows = json.loads(resp.read())
            all_rows.extend(rows)
            if len(rows) < page_size:
                break
            offset += page_size
        return all_rows

    def post(self, table, rows, upsert=False, on_conflict=None):
        """POST (insert) rows into a table. Returns count inserted."""
        url = f"{self.base_url}/rest/v1/{table}"
        if upsert and on_conflict:
            url += f"?on_conflict={on_conflict}"
        headers = {**self.headers}
        if upsert:
            headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
        data = json.dumps(rows).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        try:
            resp = urllib.request.urlopen(req, timeout=60)
            return len(rows)
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8")
            print(f"\n    REST ERROR {e.code}: {body[:300]}")
            return 0

    def rpc(self, function_name, params=None):
        """Call a PostgreSQL function via RPC."""
        url = f"{self.base_url}/rest/v1/rpc/{function_name}"
        data = json.dumps(params or {}).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=self.headers, method="POST")
        try:
            resp = urllib.request.urlopen(req, timeout=120)
            return resp.read().decode("utf-8")
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8")
            print(f"\n    RPC ERROR {e.code}: {body[:300]}")
            return None

    def batch_insert(self, table, rows, label="rows", upsert=False, on_conflict=None):
        """Insert rows in batches with progress reporting."""
        total = len(rows)
        loaded = 0
        start = time.time()
        for i in range(0, total, BATCH_SIZE):
            batch = rows[i:i + BATCH_SIZE]
            count = self.post(table, batch, upsert=upsert, on_conflict=on_conflict)
            loaded += count
            elapsed = time.time() - start
            rate = loaded / elapsed if elapsed > 0 else 0
            print(f"    {loaded:,}/{total:,} {label} ({rate:.0f}/sec)...", end="\r")
        elapsed = time.time() - start
        print(f"    {loaded:,}/{total:,} {label} in {elapsed:.1f}s                    ")
        return loaded


# ============================================================
# HELPERS
# ============================================================

def norm_discipline(name):
    if not name:
        return None, None
    name = name.strip()
    if name.lower().startswith("weight throw"):
        return None, None
    weight = None
    base_name = name
    match = re.match(r'^(.+?)\s*\((.+?)\)\s*$', name)
    if match:
        base_name = match.group(1).strip()
        weight_str = match.group(2).strip()
        if weight_str.lower() == "old":
            return None, None
        weight_str = weight_str.split("/")[0]
        weight_str = weight_str.lower().replace("gr", "g").replace("kg", "").strip()
        if weight_str.endswith("g"):
            try:
                weight = float(weight_str.rstrip("g").replace(",", ".")) / 1000
            except ValueError:
                weight = None
        else:
            try:
                weight = float(weight_str.replace(",", "."))
            except ValueError:
                weight = None
    normalised = DISCIPLINE_NORMALIZE.get(base_name.lower(), base_name)
    return normalised, weight


def norm_gender(g):
    if not g:
        return "Male"
    return GENDER_MAP.get(g.strip(), "Male")


def to_date_str(val):
    """Convert to ISO date string for JSON."""
    if val is None:
        return None
    from datetime import datetime, date
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%d %b %Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def to_float(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip().rstrip("m").strip()
    try:
        v = float(val)
        return v if v == v else None  # NaN check
    except (ValueError, TypeError):
        return None


def to_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def to_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.lower() in ("true", "1", "yes")
    return False


# ============================================================
# THROWS MIGRATION
# ============================================================

def migrate_throws(api, disc_lookup, athlete_lookup):
    if not os.path.exists(THROWS_EXCEL):
        print(f"  ERROR: Throws Excel not found at {THROWS_EXCEL}")
        return

    print(f"\nLoading throws workbook: {THROWS_EXCEL}")
    wb = openpyxl.load_workbook(THROWS_EXCEL, read_only=True, data_only=True)

    try:
        # ── Step 1: Load Athletes ──
        print("\n  --- Loading Throws Athletes ---")
        ws = wb["Athlete Summary"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        new_athletes = []
        existing = 0
        for row in rows:
            name = row[0]
            if not name:
                continue
            name = name.strip()
            if name in athlete_lookup:
                existing += 1
                continue

            gender = "M" if norm_gender(row[1]) == "Male" else "F"
            dob = to_date_str(row[2])
            country = row[3]
            doping = row[5] if len(row) > 5 else None

            new_athletes.append({
                "name": name, "gender": gender, "country": country,
                "date_of_birth": dob, "is_olympic": True, "doping_flag": doping
            })

        print(f"    {existing} already in DB, {len(new_athletes)} new athletes")

        if new_athletes:
            api.batch_insert("athletes", new_athletes, "athletes")
            # Refresh lookup
            all_athletes = api.get("athletes", "select=id,name")
            athlete_lookup.clear()
            for a in all_athletes:
                athlete_lookup[a["name"]] = a["id"]
            print(f"    Total athletes now: {len(athlete_lookup)}")

        # ── Step 2: Load Personal Bests ──
        print("\n  --- Loading Throws Personal Bests ---")
        ws = wb["Personal Bests"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        pb_data = []
        skipped = 0
        for row in rows:
            name = row[0]
            if not name:
                skipped += 1
                continue
            name = name.strip()
            athlete_id = athlete_lookup.get(name)
            if not athlete_id:
                skipped += 1
                continue

            disc_name, impl_weight = norm_discipline(row[1])
            if not disc_name:
                skipped += 1
                continue
            gender = norm_gender(row[2])
            disc_code = THROWS_DISCIPLINE_MAP.get((disc_name, gender))
            if not disc_code or disc_code not in disc_lookup:
                skipped += 1
                continue

            pb_dist = to_float(row[3])
            if not pb_dist or pb_dist <= 0:
                skipped += 1
                continue

            pb_data.append({
                "athlete_id": athlete_id,
                "discipline_id": disc_lookup[disc_code],
                "pb_time": pb_dist,
                "pb_date": to_date_str(row[4]),
                "total_races": to_int(row[5]),
                "avg_time": to_float(row[6]),
                "median_time": to_float(row[7]),
                "std_dev": to_float(row[8]),
            })

        print(f"    {len(pb_data)} PBs to load (skipped {skipped})")
        if pb_data:
            api.batch_insert("personal_bests", pb_data, "personal bests", upsert=True, on_conflict="athlete_id,discipline_id")

        # ── Step 3: Load Career Throws ──
        print("\n  --- Loading Career Throws ---")
        ws = wb["Career Throws"]

        race_rows = []
        total_skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            athlete_name = row[9]
            if not athlete_name:
                total_skipped += 1
                continue
            athlete_name = athlete_name.strip()
            athlete_id = athlete_lookup.get(athlete_name)
            if not athlete_id:
                total_skipped += 1
                continue

            disc_name, impl_weight = norm_discipline(row[1])
            if not disc_name:
                total_skipped += 1
                continue
            gender = norm_gender(row[14])
            disc_code = THROWS_DISCIPLINE_MAP.get((disc_name, gender))
            if not disc_code or disc_code not in disc_lookup:
                total_skipped += 1
                continue

            distance = to_float(row[2])
            if not distance or distance <= 0:
                total_skipped += 1
                continue

            race_rows.append({
                "athlete_id": athlete_id,
                "discipline_id": disc_lookup[disc_code],
                "race_date": to_date_str(row[0]),
                "time_seconds": distance,
                "wind_mps": None,
                "wind_legal": None,
                "competition": str(row[4])[:300] if row[4] else None,
                "country": None,
                "category": str(row[5])[:50] if row[5] else None,
                "race_type": str(row[6])[:50] if row[6] else None,
                "place": str(row[7])[:10] if row[7] else None,
                "score": to_int(row[8]),
                "age_decimal": to_float(row[11]),
                "age_years": to_int(row[12]),
                "is_olympic_race": to_bool(row[13]),
                "season_year": to_int(row[10]),
            })

        print(f"    {len(race_rows):,} rows to load (skipped {total_skipped:,})")
        if race_rows:
            api.batch_insert("race_results", race_rows, "career throws")

    finally:
        wb.close()


# ============================================================
# JUMPS MIGRATION
# ============================================================

def migrate_jumps(api, disc_lookup, athlete_lookup):
    if not os.path.isdir(JUMPS_DIR):
        print(f"  ERROR: Jumps directory not found at {JUMPS_DIR}")
        return

    # ── Step 1: Collect all unique athletes ──
    print("\n  --- Scanning Jumps Athletes ---")
    athlete_set = {}

    for csv_file, (disc_name, gender) in JUMPS_CSV_MAP.items():
        filepath = os.path.join(JUMPS_DIR, csv_file)
        if not os.path.exists(filepath):
            continue
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("athlete_name", "").strip()
                if not name or name in athlete_set:
                    continue
                g = "M" if norm_gender(row.get("Gender", gender)) == "Male" else "F"
                athlete_set[name] = {
                    "name": name, "gender": g,
                    "country": row.get("Nationality", ""),
                    "date_of_birth": to_date_str(row.get("DOB", "")),
                    "is_olympic": False, "doping_flag": None
                }

    new_athletes = [v for k, v in athlete_set.items() if k not in athlete_lookup]
    existing = len(athlete_set) - len(new_athletes)
    print(f"    {existing} already in DB, {len(new_athletes)} new athletes")

    if new_athletes:
        api.batch_insert("athletes", new_athletes, "athletes")
        all_athletes = api.get("athletes", "select=id,name")
        athlete_lookup.clear()
        for a in all_athletes:
            athlete_lookup[a["name"]] = a["id"]
        print(f"    Total athletes now: {len(athlete_lookup)}")

    # ── Step 2: Compute and load Personal Bests ──
    print("\n  --- Computing Jumps Personal Bests ---")
    pb_tracker = {}

    for csv_file, (disc_name, gender) in JUMPS_CSV_MAP.items():
        filepath = os.path.join(JUMPS_DIR, csv_file)
        if not os.path.exists(filepath):
            continue
        disc_code = JUMPS_DISCIPLINE_MAP.get((disc_name, norm_gender(gender)))
        if not disc_code:
            continue
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("athlete_name", "").strip()
                perf = to_float(row.get("Performance", ""))
                if not name or not perf or perf <= 0:
                    continue
                key = (name, disc_code)
                if key not in pb_tracker:
                    pb_tracker[key] = {"best": perf, "date": to_date_str(row.get("Date", "")),
                                       "count": 0, "sum": 0, "values": []}
                entry = pb_tracker[key]
                entry["count"] += 1
                entry["sum"] += perf
                entry["values"].append(perf)
                if perf > entry["best"]:
                    entry["best"] = perf
                    entry["date"] = to_date_str(row.get("Date", ""))

    pb_data = []
    skipped = 0
    for (name, disc_code), info in pb_tracker.items():
        athlete_id = athlete_lookup.get(name)
        if not athlete_id or disc_code not in disc_lookup:
            skipped += 1
            continue
        avg_val = info["sum"] / info["count"] if info["count"] else None
        vals = sorted(info["values"])
        median_val = vals[len(vals) // 2] if vals else None
        std_dev = None
        if len(vals) > 1 and avg_val:
            std_dev = round(math.sqrt(sum((v - avg_val) ** 2 for v in vals) / (len(vals) - 1)), 4)

        pb_data.append({
            "athlete_id": athlete_id,
            "discipline_id": disc_lookup[disc_code],
            "pb_time": info["best"],
            "pb_date": info["date"],
            "total_races": info["count"],
            "avg_time": round(avg_val, 3) if avg_val else None,
            "median_time": round(median_val, 3) if median_val else None,
            "std_dev": std_dev,
        })

    print(f"    {len(pb_data)} PBs computed (skipped {skipped})")
    if pb_data:
        api.batch_insert("personal_bests", pb_data, "personal bests", upsert=True, on_conflict="athlete_id,discipline_id")

    # ── Step 3: Load Race Results ──
    print("\n  --- Loading Jumps Race Results ---")
    grand_total = 0

    for csv_file, (disc_name, gender) in JUMPS_CSV_MAP.items():
        filepath = os.path.join(JUMPS_DIR, csv_file)
        if not os.path.exists(filepath):
            continue
        disc_code = JUMPS_DISCIPLINE_MAP.get((disc_name, norm_gender(gender)))
        if not disc_code or disc_code not in disc_lookup:
            continue
        disc_id = disc_lookup[disc_code]

        race_rows = []
        file_skipped = 0

        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("athlete_name", "").strip()
                if not name:
                    file_skipped += 1
                    continue
                athlete_id = athlete_lookup.get(name)
                if not athlete_id:
                    file_skipped += 1
                    continue
                perf = to_float(row.get("Performance", ""))
                if not perf or perf <= 0:
                    file_skipped += 1
                    continue

                wind = to_float(row.get("Wind", ""))
                wind_legal = True if wind is not None and wind <= 2.0 else (None if wind is None else False)
                category_val = (row.get("Category", "") or "")[:50] or None
                is_olympic = "OW" in (category_val or "") or "OG" in (category_val or "")

                race_rows.append({
                    "athlete_id": athlete_id,
                    "discipline_id": disc_id,
                    "race_date": to_date_str(row.get("Date", "")),
                    "time_seconds": perf,
                    "wind_mps": wind,
                    "wind_legal": wind_legal,
                    "competition": (row.get("Competition", "") or "")[:300] or None,
                    "country": None,
                    "category": category_val,
                    "race_type": (row.get("Race", "") or "")[:50] or None,
                    "place": (row.get("Place", "") or "")[:10] or None,
                    "score": to_int(row.get("Score", "")),
                    "age_decimal": to_float(row.get("Age_At_Event_Decimal", "")),
                    "age_years": to_int(row.get("Age_At_Event_Years", "")),
                    "is_olympic_race": is_olympic,
                    "season_year": to_int(row.get("Year", "")),
                })

        print(f"\n    {csv_file}: {len(race_rows):,} rows (skipped {file_skipped})")
        if race_rows:
            api.batch_insert("race_results", race_rows, f"{disc_code} results")
            grand_total += len(race_rows)

    print(f"\n    Total jumps loaded: {grand_total:,}")


# ============================================================
# MAIN
# ============================================================

def run_migration(supabase_url, service_key, throws_only=False, jumps_only=False):
    api = SupabaseREST(supabase_url, service_key)

    # ── Test connection ──
    print("Testing connection...")
    try:
        disciplines = api.get("disciplines", "select=id,code&order=code")
        print(f"  Connected! {len(disciplines)} disciplines in DB")
    except Exception as e:
        print(f"  ERROR: {e}")
        sys.exit(1)

    # ── Register jumps disciplines ──
    print("\n--- Registering Jumps Disciplines ---")
    existing_codes = {d["code"] for d in disciplines}
    new_discs = [d for d in JUMPS_DISCIPLINES_INSERT if d["code"] not in existing_codes]
    if new_discs:
        api.post("disciplines", new_discs)
        print(f"  Added {len(new_discs)} jumps disciplines")
    else:
        print("  All jumps disciplines already exist")

    # Refresh discipline lookup
    disciplines = api.get("disciplines", "select=id,code&order=code")
    disc_lookup = {d["code"]: d["id"] for d in disciplines}
    print(f"  Disciplines: {sorted(disc_lookup.keys())}")

    # ── Load athlete lookup ──
    print("\n--- Loading Athlete Lookup ---")
    all_athletes = api.get("athletes", "select=id,name")
    athlete_lookup = {a["name"]: a["id"] for a in all_athletes}
    print(f"  {len(athlete_lookup)} existing athletes")

    # ── Run migrations ──
    if not jumps_only:
        print("\n" + "=" * 60)
        print("THROWS MIGRATION")
        print("=" * 60)
        migrate_throws(api, disc_lookup, athlete_lookup)

    if not throws_only:
        print("\n" + "=" * 60)
        print("JUMPS MIGRATION")
        print("=" * 60)
        migrate_jumps(api, disc_lookup, athlete_lookup)

    # ── Populate season_bests ──
    print("\n" + "=" * 60)
    print("POPULATING SEASON BESTS")
    print("=" * 60)
    print("  Running populate_season_bests()...")
    result = api.rpc("populate_season_bests")
    if result is not None:
        print("  Done!")
    else:
        print("  WARNING: RPC call may have failed. Check Supabase logs.")

    # ── Final stats ──
    print("\n" + "=" * 60)
    print("FINAL STATS")
    print("=" * 60)
    for table in ["disciplines", "athletes", "race_results", "personal_bests", "season_bests"]:
        try:
            # Use HEAD request with count
            url = f"{supabase_url}/rest/v1/{table}?select=id"
            req = urllib.request.Request(url, headers={
                "apikey": service_key,
                "Authorization": f"Bearer {service_key}",
                "Prefer": "count=exact",
                "Range": "0-0",
            })
            resp = urllib.request.urlopen(req, timeout=30)
            content_range = resp.headers.get("Content-Range", "")
            # Format: "0-0/12345"
            total = content_range.split("/")[-1] if "/" in content_range else "?"
            print(f"  {table}: {total} rows")
        except Exception:
            print(f"  {table}: (could not count)")

    print("\nDone!")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate throws & jumps via Supabase REST API")
    parser.add_argument("--supabase-url", required=True, help="Supabase project URL")
    parser.add_argument("--service-key", required=True, help="Supabase service_role key")
    parser.add_argument("--throws-only", action="store_true")
    parser.add_argument("--jumps-only", action="store_true")
    args = parser.parse_args()

    run_migration(args.supabase_url, args.service_key,
                  throws_only=args.throws_only, jumps_only=args.jumps_only)
